import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from typing import Dict, Any

def generate_costing_excel(costing_data: Dict[str, Any], output_path: str) -> str:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Costing Sheet Summary"
    
    # Enable grid lines explicitly
    ws.views.sheetView[0].showGridLines = True
    
    # Styles Definition (Curated Premium Palette - Corporate Dark Slate and Mint accents)
    font_family = "Segoe UI"
    
    title_font = Font(name=font_family, size=16, bold=True, color="FFFFFF")
    section_font = Font(name=font_family, size=11, bold=True, color="1F2937")
    header_font = Font(name=font_family, size=11, bold=True, color="FFFFFF")
    bold_font = Font(name=font_family, size=11, bold=True, color="1F2937")
    regular_font = Font(name=font_family, size=11, color="374151")
    note_font = Font(name=font_family, size=9, italic=True, color="6B7280")
    
    title_fill = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid") # Dark Slate Gray
    header_fill = PatternFill(start_color="4B5563", end_color="4B5563", fill_type="solid") # Medium Slate
    total_fill = PatternFill(start_color="E5E7EB", end_color="E5E7EB", fill_type="solid") # Light Slate Gray Accent
    fob_fill = PatternFill(start_color="D1FAE5", end_color="D1FAE5", fill_type="solid") # Soft Light Green (Mint) for FOB price
    
    thin_border_side = Side(style='thin', color='D1D5DB')
    double_border_side = Side(style='double', color='1F2937')
    
    thin_border = Border(left=thin_border_side, right=thin_border_side, top=thin_border_side, bottom=thin_border_side)
    total_border = Border(top=thin_border_side, bottom=double_border_side)
    
    align_left = Alignment(horizontal='left', vertical='center')
    align_right = Alignment(horizontal='right', vertical='center')
    align_center = Alignment(horizontal='center', vertical='center')
    
    # 1. Title Block (Merged A1:E1)
    ws.merge_cells('A1:E2')
    title_cell = ws['A1']
    title_cell.value = "  GARMENT TECH PACK COSTING RESPONSE"
    title_cell.font = title_font
    title_cell.fill = title_fill
    title_cell.alignment = Alignment(horizontal='left', vertical='center')
    
    # Apply fill to the rest of the merged title block cells to ensure it looks styled
    for row in range(1, 3):
        for col in range(1, 6):
            ws.cell(row=row, column=col).fill = title_fill
            
    # 2. Metadata Block (Rows 4-7)
    metadata = [
        ("Style Name:", costing_data.get("style_name")),
        ("Style Number:", costing_data.get("style_number")),
        ("Costing Date:", costing_data.get("costing_date")),
        ("Currency:", costing_data.get("currency", "USD"))
    ]
    
    curr_row = 4
    for label, val in metadata:
        ws.cell(row=curr_row, column=1, value=label).font = bold_font
        ws.cell(row=curr_row, column=2, value=val).font = regular_font
        ws.cell(row=curr_row, column=1).alignment = align_left
        ws.cell(row=curr_row, column=2).alignment = align_left
        curr_row += 1
        
    # Fabric consumption summary metadata
    fab_cons = costing_data.get("fabric_consumption", {})
    ws.cell(row=4, column=4, value="Est. Fabric Consumption:").font = bold_font
    ws.cell(row=4, column=5, value=f"{fab_cons.get('main_fabric_yards_per_piece', 0.0)} yards/pc").font = regular_font
    ws.cell(row=5, column=4, value="Fabric Width:").font = bold_font
    ws.cell(row=5, column=5, value=f"{fab_cons.get('fabric_width_inches', 60.0)} inches").font = regular_font
    ws.cell(row=6, column=4, value="Garment Type:").font = bold_font
    ws.cell(row=6, column=5, value=f"{fab_cons.get('garment_type_factor', 'T-shirt')}").font = regular_font
    
    for r in range(4, 7):
        ws.cell(row=r, column=4).alignment = align_left
        ws.cell(row=r, column=5).alignment = align_left

    # 3. Bill of Materials (BOM) Table (Starts row 9)
    table_start_row = 9
    headers = ["Item Description", "Consumption", "Unit", "Unit Rate ($)", "Cost per Piece ($)"]
    
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=table_start_row, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = align_center if col_idx in [2, 3] else (align_right if col_idx >= 4 else align_left)
        cell.border = thin_border
        
    curr_row = table_start_row + 1
    material_breakdown = costing_data.get("material_cost_breakdown", [])
    
    for item in material_breakdown:
        c1 = ws.cell(row=curr_row, column=1, value=item.get("item"))
        c2 = ws.cell(row=curr_row, column=2, value=item.get("consumption"))
        c3 = ws.cell(row=curr_row, column=3, value=item.get("unit"))
        c4 = ws.cell(row=curr_row, column=4, value=item.get("rate"))
        c5 = ws.cell(row=curr_row, column=5, value=item.get("cost"))
        
        # Formatting
        c1.alignment = align_left
        c2.alignment = align_center
        c3.alignment = align_center
        c4.alignment = align_right
        c5.alignment = align_right
        
        c4.number_format = "$#,##0.00"
        c5.number_format = "$#,##0.00"
        c2.number_format = "0.00"
        
        for c in [c1, c2, c3, c4, c5]:
            c.font = regular_font
            c.border = thin_border
        curr_row += 1
        
    # 4. Summary Costing Calculations (CMT, Factory, FOB)
    summary_items = [
        ("Total Material Cost", costing_data.get("total_material_cost"), total_fill, total_border),
        ("CMT Cost (Cut-Make-Trim)", costing_data.get("cmt_cost"), None, thin_border),
        ("Factory Cost", costing_data.get("factory_cost_per_piece"), total_fill, total_border),
        ("Markup Margin ({}%)".format(costing_data.get("markup_percentage", 15)), 
         round(costing_data.get("factory_cost_per_piece", 0.0) * (costing_data.get("markup_percentage", 15)/100), 2), None, thin_border),
        ("Final FOB Price per Piece", costing_data.get("fob_price_per_piece"), fob_fill, Border(top=Side(style='thin', color='1F2937'), bottom=double_border_side))
    ]
    
    curr_row += 1
    for label, val, fill, border in summary_items:
        ws.merge_cells(start_row=curr_row, start_column=1, end_row=curr_row, end_column=4)
        lbl_cell = ws.cell(row=curr_row, column=1, value=label)
        lbl_cell.font = bold_font
        lbl_cell.alignment = align_right
        
        val_cell = ws.cell(row=curr_row, column=5, value=val)
        val_cell.font = bold_font
        val_cell.alignment = align_right
        val_cell.number_format = "$#,##0.00"
        
        # Border and fills application for merged area
        for col_idx in range(1, 6):
            c = ws.cell(row=curr_row, column=col_idx)
            if fill:
                c.fill = fill
            if border:
                # Apply specific border properties
                c.border = Border(
                    top=border.top if border.top else Side(style=None),
                    bottom=border.bottom if border.bottom else Side(style=None),
                    left=thin_border_side if col_idx == 1 else (Side(style=None)),
                    right=thin_border_side if col_idx == 5 else (Side(style=None))
                )
                
        val_cell.border = Border(top=border.top, bottom=border.bottom, right=thin_border_side)
        curr_row += 1
        
    # 5. Order Volume Block (Row space)
    curr_row += 1
    ws.cell(row=curr_row, column=1, value="ORDER QUANTITY SUMMARY").font = section_font
    curr_row += 1
    
    qty_data = costing_data.get("order_quantity", {})
    sizes_listed = [k for k in qty_data.keys() if k != "total"]
    
    # Headers
    ws.cell(row=curr_row, column=1, value="Size").font = header_font
    ws.cell(row=curr_row, column=1).fill = header_fill
    ws.cell(row=curr_row, column=1).alignment = align_center
    ws.cell(row=curr_row, column=1).border = thin_border
    
    ws.cell(row=curr_row, column=2, value="Quantity (pcs)").font = header_font
    ws.cell(row=curr_row, column=2).fill = header_fill
    ws.cell(row=curr_row, column=2).alignment = align_center
    ws.cell(row=curr_row, column=2).border = thin_border
    
    qty_row_start = curr_row + 1
    curr_row += 1
    
    for size in sizes_listed:
        s_cell = ws.cell(row=curr_row, column=1, value=size)
        q_cell = ws.cell(row=curr_row, column=2, value=qty_data[size])
        
        s_cell.font = regular_font
        s_cell.alignment = align_center
        s_cell.border = thin_border
        
        q_cell.font = regular_font
        q_cell.alignment = align_right
        q_cell.number_format = "#,##0"
        q_cell.border = thin_border
        curr_row += 1
        
    # Total Order Qty Row
    ws.cell(row=curr_row, column=1, value="Total Quantity").font = bold_font
    ws.cell(row=curr_row, column=1).alignment = align_center
    ws.cell(row=curr_row, column=1).border = total_border
    ws.cell(row=curr_row, column=1).fill = total_fill
    
    tot_q_cell = ws.cell(row=curr_row, column=2, value=qty_data.get("total", 0))
    tot_q_cell.font = bold_font
    tot_q_cell.alignment = align_right
    tot_q_cell.number_format = "#,##0"
    tot_q_cell.border = total_border
    tot_q_cell.fill = total_fill
    curr_row += 1
    
    # Total Contract Value
    ws.cell(row=curr_row, column=1, value="Total Order Value").font = bold_font
    ws.cell(row=curr_row, column=1).alignment = align_center
    ws.cell(row=curr_row, column=1).border = total_border
    ws.cell(row=curr_row, column=1).fill = fob_fill
    
    tot_val_cell = ws.cell(row=curr_row, column=2, value=costing_data.get("total_order_value", 0.0))
    tot_val_cell.font = bold_font
    tot_val_cell.alignment = align_right
    tot_val_cell.number_format = "$#,##0.00"
    tot_val_cell.border = total_border
    tot_val_cell.fill = fob_fill
    curr_row += 2
    
    # 6. Notes Block
    ws.cell(row=curr_row, column=1, value="EXPORTER NOTES:").font = section_font
    curr_row += 1
    
    for note in costing_data.get("notes", []):
        ws.merge_cells(start_row=curr_row, start_column=1, end_row=curr_row, end_column=5)
        note_cell = ws.cell(row=curr_row, column=1, value=f"- {note}")
        note_cell.font = note_font
        note_cell.alignment = align_left
        curr_row += 1
        
    # Auto-adjust column widths
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        
        # Don't auto-size based on title block or note blocks because they are merged and very wide
        for cell in col:
            val = str(cell.value or '')
            if cell.row in [1, 2] or cell.row > curr_row - len(costing_data.get("notes", [])) - 2:
                continue
            if len(val) > max_len:
                max_len = len(val)
                
        ws.column_dimensions[col_letter].width = max(max_len + 4, 15)
        
    # Explicitly set Item column to be wider
    ws.column_dimensions['A'].width = 30
    
    wb.save(output_path)
    return output_path
